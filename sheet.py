from openpyxl import workbook

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.xml.constants import MAX_COLUMN, MAX_ROW, MIN_COLUMN, MIN_ROW


workbook = xl.load_workbook("transactions.xlsx")

sheet = workbook["Sheet1"]


for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)

    corrected_price = float(cell.value) * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


data = Reference(sheet, min_row=2, max_row=sheet.max_row,
                 min_col=4, max_col=4)
chart = BarChart()
chart.title = "Bar Chart"
chart.y_axis.title = "Price"
chart.x_axis.title = "Product"
chart.add_data(data)
sheet.add_chart(chart, "E2")


sheet2 = workbook["Sheet2"]

data1 = Reference(sheet2, min_row=1, max_row=sheet2.max_row,
                  min_col=2, max_col=3)
cat = Reference(sheet2, min_col=1, min_row=2, max_row=4)

chart1 = BarChart()
chart1.x_axis.title = "Year"
chart1.y_axis.title = "Price"
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cat)
sheet2.add_chart(chart1, "E4")


workbook.save("transactions2.xlsx")
